import os
import csv
import json
import logging
from typing import Dict, List, Any
from datetime import datetime
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.pdfgen import canvas
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

logger = logging.getLogger(__name__)

class ReportGenerator:
    """Generates reports in different formats"""
    
    def __init__(self):
        self.temp_folder = os.path.join(os.getcwd(), 'temp')
        os.makedirs(self.temp_folder, exist_ok=True)
    
    def generate_pdf(self, data: Dict[str, Any], session_id: str) -> str:
        """Generate PDF report"""
        try:
            filename = f"pedido_compra_{session_id}.pdf"
            filepath = os.path.join(self.temp_folder, filename)
            
            # Create PDF document
            doc = SimpleDocTemplate(filepath, pagesize=A4)
            story = []
            
            # Styles
            styles = getSampleStyleSheet()
            title_style = ParagraphStyle(
                'CustomTitle',
                parent=styles['Heading1'],
                fontSize=16,
                spaceAfter=30,
                alignment=1  # Center alignment
            )
            
            # Title
            title = Paragraph("PEDIDO DE COMPRA CONSOLIDADO", title_style)
            story.append(title)
            story.append(Spacer(1, 20))
            
            # Summary info
            summary_data = [
                ["Data de Geração:", datetime.now().strftime("%d/%m/%Y %H:%M:%S")],
                ["Total de Produtos:", str(data['total_products'])],
                ["Valor Total:", f"R$ {data['total_value']:.2f}"],
                ["Arquivos Processados:", str(data['processing_info']['processed_files'])],
                ["Produtos Extraídos:", str(data['processing_info']['extracted_products'])]
            ]
            
            summary_table = Table(summary_data, colWidths=[2*inch, 3*inch])
            summary_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (0, -1), colors.lightgrey),
                ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
                ('FONTSIZE', (0, 0), (-1, -1), 10),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 12),
                ('BACKGROUND', (1, 0), (1, -1), colors.white),
                ('GRID', (0, 0), (-1, -1), 1, colors.black)
            ]))
            
            story.append(summary_table)
            story.append(Spacer(1, 30))
            
            # Products table
            products_title = Paragraph("PRODUTOS CONSOLIDADOS", styles['Heading2'])
            story.append(products_title)
            story.append(Spacer(1, 10))
            
            # Table headers
            headers = ['Código', 'Referência', 'Descrição', 'Qtd.', 'Valor Unit.', 'Valor Total', 'Fonte']
            table_data = [headers]
            
            # Add products
            for product in data['products']:
                row = [
                    product['codigo'] or '-',
                    product['referencia'] or '-',
                    product['descricao'][:50] + '...' if len(product['descricao']) > 50 else product['descricao'],
                    str(int(product['quantidade'])),
                    f"R$ {product['valor_unitario']:.2f}",
                    f"R$ {product['valor_total']:.2f}",
                    product['fonte'][:20] + '...' if len(product['fonte']) > 20 else product['fonte']
                ]
                table_data.append(row)
            
            # Create table
            products_table = Table(table_data, colWidths=[0.8*inch, 0.8*inch, 2.2*inch, 0.5*inch, 0.8*inch, 0.8*inch, 1.1*inch])
            products_table.setStyle(TableStyle([
                # Header style
                ('BACKGROUND', (0, 0), (-1, 0), colors.darkblue),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 9),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                
                # Data style
                ('BACKGROUND', (0, 1), (-1, -1), colors.white),
                ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
                ('ALIGN', (0, 1), (2, -1), 'LEFT'),  # Left align text columns
                ('ALIGN', (3, 1), (-1, -1), 'RIGHT'), # Right align number columns
                ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
                ('FONTSIZE', (0, 1), (-1, -1), 8),
                ('BOTTOMPADDING', (0, 1), (-1, -1), 6),
                
                # Grid
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
                
                # Zebra stripes
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey])
            ]))
            
            story.append(products_table)
            
            # Build PDF
            doc.build(story)
            
            logger.info(f"PDF generated: {filepath}")
            return filepath
            
        except Exception as e:
            logger.error(f"Error generating PDF: {str(e)}")
            raise e
    
    def generate_excel(self, data: Dict[str, Any], session_id: str) -> str:
        """Generate Excel report"""
        try:
            filename = f"pedido_compra_{session_id}.xlsx"
            filepath = os.path.join(self.temp_folder, filename)
            
            wb = Workbook()
            ws = wb.active
            ws.title = "Pedido de Compra"
            
            # Styles
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            center_alignment = Alignment(horizontal="center", vertical="center")
            
            # Title
            ws.merge_cells("A1:G1")
            ws["A1"] = "PEDIDO DE COMPRA CONSOLIDADO"
            ws["A1"].font = Font(bold=True, size=16)
            ws["A1"].alignment = center_alignment
            
            # Summary
            ws["A3"] = "Data de Geração:"
            ws["B3"] = datetime.now().strftime("%d/%m/%Y %H:%M:%S")
            ws["A4"] = "Total de Produtos:"
            ws["B4"] = data['total_products']
            ws["A5"] = "Valor Total:"
            ws["B5"] = f"R$ {data['total_value']:.2f}"
            ws["A6"] = "Arquivos Processados:"
            ws["B6"] = data['processing_info']['processed_files']
            ws["A7"] = "Produtos Extraídos:"
            ws["B7"] = data['processing_info']['extracted_products']
            
            # Headers
            headers = ['Código', 'Referência', 'Descrição', 'Quantidade', 'Valor Unitário', 'Valor Total', 'Fonte']
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=9, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = center_alignment
            
            # Data
            for row, product in enumerate(data['products'], 10):
                ws.cell(row=row, column=1, value=product['codigo'] or '-')
                ws.cell(row=row, column=2, value=product['referencia'] or '-')
                ws.cell(row=row, column=3, value=product['descricao'])
                ws.cell(row=row, column=4, value=int(product['quantidade']))
                ws.cell(row=row, column=5, value=product['valor_unitario'])
                ws.cell(row=row, column=6, value=product['valor_total'])
                ws.cell(row=row, column=7, value=product['fonte'])
            
            # Adjust column widths
            ws.column_dimensions['A'].width = 15
            ws.column_dimensions['B'].width = 15
            ws.column_dimensions['C'].width = 40
            ws.column_dimensions['D'].width = 12
            ws.column_dimensions['E'].width = 15
            ws.column_dimensions['F'].width = 15
            ws.column_dimensions['G'].width = 25
            
            # Save file
            wb.save(filepath)
            
            logger.info(f"Excel generated: {filepath}")
            return filepath
            
        except Exception as e:
            logger.error(f"Error generating Excel: {str(e)}")
            raise e
    
    def generate_csv(self, data: Dict[str, Any], session_id: str) -> str:
        """Generate CSV report"""
        try:
            filename = f"pedido_compra_{session_id}.csv"
            filepath = os.path.join(self.temp_folder, filename)
            
            with open(filepath, 'w', newline='', encoding='utf-8') as csvfile:
                writer = csv.writer(csvfile)
                
                # Header info
                writer.writerow(['PEDIDO DE COMPRA CONSOLIDADO'])
                writer.writerow([])
                writer.writerow(['Data de Geração:', datetime.now().strftime("%d/%m/%Y %H:%M:%S")])
                writer.writerow(['Total de Produtos:', data['total_products']])
                writer.writerow(['Valor Total:', f"R$ {data['total_value']:.2f}"])
                writer.writerow(['Arquivos Processados:', data['processing_info']['processed_files']])
                writer.writerow(['Produtos Extraídos:', data['processing_info']['extracted_products']])
                writer.writerow([])
                
                # Headers
                headers = ['Código', 'Referência', 'Descrição', 'Quantidade', 'Valor Unitário', 'Valor Total', 'Fonte']
                writer.writerow(headers)
                
                # Data
                for product in data['products']:
                    row = [
                        product['codigo'] or '-',
                        product['referencia'] or '-',
                        product['descricao'],
                        int(product['quantidade']),
                        f"{product['valor_unitario']:.2f}",
                        f"{product['valor_total']:.2f}",
                        product['fonte']
                    ]
                    writer.writerow(row)
            
            logger.info(f"CSV generated: {filepath}")
            return filepath
            
        except Exception as e:
            logger.error(f"Error generating CSV: {str(e)}")
            raise e
